from flask.app import Flask
from flask.helpers import flash
import openpyxl
from openpyxl.descriptors.base import Length
import os
from flask import Blueprint , render_template,request
import glob

def Filelistfun() :
    fileslist=glob.glob("website\ExcelFiles/**/*.xlsx", recursive=True)
    fileslistname=[]
    for f in fileslist: 
        fileslistname.append(str(f).replace('website\ExcelFiles\\',''))
    return fileslistname

search=Blueprint('search', __name__)
@search.route('/search' , methods=['GET', 'POST'])
def searching():
    return render_template("search.html")


@search.route('/result',methods=['GET', 'POST'])
def result():
    if request.method=='POST':
        try :
            filename=request.form['F_inside_result']
            ps =openpyxl.load_workbook(str( 'website\ExcelFiles' +'/'+str(filename) +'.xlsx'))
            sh= ps['Feuil1']
            sh['C'+'3'].value = request.form.get('MLOtype')
            ps.save(str('website\ExcelFiles' +'/'+filename+'.xlsx'))
            return render_template("result.html" , sh=sh , fnx=filename)
        except:
            try:
                filename=request.form['F_inside_done']
                fileNameNoXlsx=filename.replace('.xlsx','')
                f ='website/ExcelFiles/'+ filename
                ps=openpyxl.load_workbook(str(f))
                sh= ps['Feuil1']
                return render_template("result.html" , sh=sh , fnx=fileNameNoXlsx)
            except :
                pass
        
    if request.method == 'POST': 
        try  :
            fn =(request.form['button'])
            fileNameNoXlsx=fn.replace('.xlsx','')
            f ='website/ExcelFiles/'+ fn
            ps =openpyxl.load_workbook(str(f))
            sh= ps['Feuil1']
            return render_template("result.html" , sh=sh , fnx=fileNameNoXlsx)
        except :
            try :
                fn =(request.form['closeAll'])
                a=True
            except :
                fn =(request.form['closeS'])
                print (request.form)
                fn=fn.split(',')[0]
                a=False
            finally :
                f ='website/ExcelFiles/'+ fn
                try :

                    os.remove(f)
                    flash ('fichier supprimé' ,category='ERR')
                except:
                    pass
                fileslistname= Filelistfun()                
                if a :
                    return render_template("all.html" ,fln=(fileslistname),fln1=(fileslistname) )
                else :
                    input=(request.form['closeS']).split(',')[1]
                    resultlist=[]
                    for f in fileslistname:
                        if input in f:
                            resultlist.append(f)
                    return render_template( "searchres.html", mloName=input, fln = (resultlist), fln1=(resultlist) )
    

@search.route('/all')
def all():
    fileslistname= Filelistfun()
    print(list.__len__(fileslistname))
    return render_template("all.html" ,fln= (fileslistname),fln1=(fileslistname) )

@search.route('/searchResult',methods=['GET', 'POST'])
def searchresult(): 
    fileslistname= Filelistfun()
    if request.method=='POST':
        input = request.form['mlo']
        resultlist=[]
        for f in fileslistname:
            if input in f:
                resultlist.append(f)
        print(list.__len__(resultlist))
        if list.__len__(resultlist) == 0:
            flash ('pas de resultat' ,category='ERR')
            return render_template("noresult.html", input=input)
        resultlist.sort()
    return render_template( "searchres.html", mloName=input, fln=(resultlist),  fln1=(resultlist)  )